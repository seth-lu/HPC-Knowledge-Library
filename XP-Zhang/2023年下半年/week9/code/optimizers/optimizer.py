import openpyxl
# 创建一个新的工作簿
workbook = openpyxl.Workbook()
sheet = workbook.active

# with open("O2_optimizers.txt", "r") as file:
# with open("O3_optimizers.txt", "r") as file:
with open("Of_optimizers .txt", "r") as file:
    row_index = 1
    for line in file:
        parts = line.strip().split(' ', 1)
        if len(parts) == 2:
            first_part = parts[0]
            second_part = parts[1].strip()
            print(first_part, " ",second_part)
            sheet.cell(row=row_index, column=1,  value=" " + first_part)
            sheet.cell(row=row_index, column=2, value=second_part)
        else:
            sheet.cell(row=row_index, column=1, value=" " + line.strip())
            print(line)
        row_index += 1

# 保存工作簿到xlsx文件
# workbook.save("O2-optimizers.xlsx")
# workbook.save("O3-optimizers.xlsx")
workbook.save("Of-optimizers.xlsx")